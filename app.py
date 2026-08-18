"""BuzzFlow — Agency OS (multi-tenant)
- Any business owner can CREATE their own workspace (becomes its one and only Owner)
- Employees JOIN a workspace via QR / invite link — never as Owner
- Every workspace's data (team, clients, tasks, updates, logo, report number)
  is completely walled off from every other workspace.
"""
import os, sqlite3, hashlib, secrets, json, io
from datetime import datetime, date, timedelta, timezone

IST = timezone(timedelta(hours=5, minutes=30))

def now_ist():
    return datetime.now(IST)

def today_ist():
    return now_ist().date()
from functools import wraps
from flask import Flask, request, jsonify, session, send_from_directory, g

DATABASE_URL = os.environ.get('DATABASE_URL', '')   # Neon/Postgres in production
USE_PG = DATABASE_URL.startswith(('postgres://', 'postgresql://'))
if USE_PG:
    import psycopg2
    import psycopg2.extras

DB = os.path.join(os.path.dirname(__file__), 'buzzflow.db')
app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))


class PgConn:
    """Small adapter so the same sqlite-style code runs on Postgres:
    translates '?' placeholders, exposes .execute/.fetchone/.fetchall/.commit,
    and returns rows accessible by column name."""
    def __init__(self, url):
        self.conn = psycopg2.connect(url)
        self.conn.autocommit = False

    class _Cur:
        def __init__(self, cur):
            self.cur = cur
        @property
        def lastrowid(self):
            row = self.cur.fetchone()
            return row['id'] if row else None
        def fetchone(self):
            return self.cur.fetchone()
        def fetchall(self):
            return self.cur.fetchall()
        def __iter__(self):
            return iter(self.cur.fetchall())

    _ID_TABLES = ('workspaces', 'users', 'clients', 'quotas', 'tasks', 'updates', 'services', 'plans', 'wpayments')

    def execute(self, sql, params=()):
        q = sql.replace('?', '%s')
        qs = q.lstrip()
        if qs.upper().startswith('INSERT') and 'RETURNING' not in q.upper():
            # only tables that actually have an id column
            target = qs.split()[2].split('(')[0].lower()
            if target in PgConn._ID_TABLES:
                q += ' RETURNING id'
        cur = self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        try:
            cur.execute(q, params)
        except Exception:
            self.conn.rollback()
            raise
        return PgConn._Cur(cur)

    def executescript(self, script):
        cur = self.conn.cursor()
        cur.execute(script)
        self.conn.commit()

    def commit(self):
        self.conn.commit()

    def close(self):
        self.conn.close()

SERVICE_TYPES = ["Reel", "Post", "Story", "Shoot", "Influencer", "Meta Ad", "Blog/Article", "Website", "SEO"]

# ---------------- db ----------------
def db():
    if 'db' not in g:
        if USE_PG:
            g.db = PgConn(DATABASE_URL)
        else:
            g.db = sqlite3.connect(DB)
            g.db.row_factory = sqlite3.Row
            g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    d = g.pop('db', None)
    if d: d.close()

def _is_unique_violation(ex):
    if isinstance(ex, sqlite3.IntegrityError):
        return True
    return ex.__class__.__name__ == 'UniqueViolation' or 'unique' in str(ex).lower()

def hashpw(pw, salt):
    return hashlib.pbkdf2_hmac('sha256', pw.encode(), salt.encode(), 100_000).hex()

SCHEMA = """
    CREATE TABLE IF NOT EXISTS workspaces(
      id {PK}, name TEXT NOT NULL, invite_code TEXT NOT NULL, created_at TEXT);
    CREATE TABLE IF NOT EXISTS users(
      id {PK}, workspace_id INTEGER NOT NULL REFERENCES workspaces(id) ON DELETE CASCADE,
      name TEXT NOT NULL, email TEXT UNIQUE NOT NULL,
      salt TEXT NOT NULL, pw TEXT NOT NULL, role TEXT NOT NULL CHECK(role IN('owner','admin','employee')),
      title TEXT DEFAULT '', active INTEGER DEFAULT 1, created_at TEXT);
    CREATE TABLE IF NOT EXISTS clients(
      id {PK}, workspace_id INTEGER NOT NULL REFERENCES workspaces(id) ON DELETE CASCADE,
      name TEXT NOT NULL, package TEXT DEFAULT '', fee BIGINT DEFAULT 0,
      contact_name TEXT DEFAULT '', contact_phone TEXT DEFAULT '', notes TEXT DEFAULT '',
      status TEXT DEFAULT 'active' CHECK(status IN('active','paused','ended')), created_at TEXT);
    CREATE TABLE IF NOT EXISTS quotas(
      id {PK}, client_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
      service TEXT NOT NULL, monthly_target INTEGER NOT NULL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS tasks(
      id {PK}, workspace_id INTEGER NOT NULL REFERENCES workspaces(id) ON DELETE CASCADE,
      title TEXT NOT NULL, client_id INTEGER REFERENCES clients(id) ON DELETE SET NULL,
      service TEXT DEFAULT 'Other', qty INTEGER DEFAULT 1,
      assignee_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
      status TEXT DEFAULT 'todo' CHECK(status IN('todo','doing','review','done')),
      priority TEXT DEFAULT 'medium' CHECK(priority IN('low','medium','high','urgent')),
      due TEXT, notes TEXT DEFAULT '', created_by INTEGER, created_at TEXT, done_at TEXT);
    CREATE TABLE IF NOT EXISTS updates(
      id {PK}, workspace_id INTEGER NOT NULL REFERENCES workspaces(id) ON DELETE CASCADE,
      user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
      client_id INTEGER REFERENCES clients(id) ON DELETE SET NULL,
      text TEXT NOT NULL, created_at TEXT);
    CREATE TABLE IF NOT EXISTS tokens(
      token TEXT PRIMARY KEY, user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
      created_at TEXT);
    CREATE TABLE IF NOT EXISTS wsettings(
      workspace_id INTEGER NOT NULL REFERENCES workspaces(id) ON DELETE CASCADE,
      key TEXT NOT NULL, value TEXT, PRIMARY KEY(workspace_id, key));
    CREATE TABLE IF NOT EXISTS services(
      id {PK}, workspace_id INTEGER NOT NULL REFERENCES workspaces(id) ON DELETE CASCADE,
      name TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS plans(
      id {PK}, name TEXT NOT NULL, price_m INTEGER DEFAULT 0, price_y INTEGER DEFAULT 0,
      max_members INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS wpayments(
      id {PK}, workspace_id INTEGER REFERENCES workspaces(id) ON DELETE CASCADE,
      ws_label TEXT DEFAULT '', amount REAL NOT NULL DEFAULT 0, note TEXT DEFAULT '', paid_on TEXT);
    CREATE TABLE IF NOT EXISTS msessions(
      token TEXT PRIMARY KEY, created_at TEXT);
    CREATE TABLE IF NOT EXISTS msettings(
      key TEXT PRIMARY KEY, value TEXT);
    CREATE TABLE IF NOT EXISTS referrals(
      id {PK}, referrer_ws INTEGER NOT NULL REFERENCES workspaces(id) ON DELETE CASCADE,
      referred_ws INTEGER REFERENCES workspaces(id) ON DELETE SET NULL,
      referred_name TEXT DEFAULT '', status TEXT DEFAULT 'pending',
      reward_note TEXT DEFAULT '', reward_days INTEGER DEFAULT 0, created_at TEXT, rewarded_at TEXT);
"""

WS_MIGRATIONS = (
    "ALTER TABLE workspaces ADD COLUMN {IFNE} plan TEXT DEFAULT 'Pro'",
    "ALTER TABLE workspaces ADD COLUMN {IFNE} expires_on TEXT",
    "ALTER TABLE workspaces ADD COLUMN {IFNE} wstatus TEXT DEFAULT 'active'",
    "ALTER TABLE workspaces ADD COLUMN {IFNE} mnotes TEXT DEFAULT ''",
    "ALTER TABLE plans ADD COLUMN {IFNE} duration_val INTEGER DEFAULT 0",
    "ALTER TABLE plans ADD COLUMN {IFNE} duration_unit TEXT DEFAULT 'days'",
    "ALTER TABLE plans ADD COLUMN {IFNE} price_pack INTEGER DEFAULT 0",
    "ALTER TABLE workspaces ADD COLUMN {IFNE} ref_code TEXT",
    "ALTER TABLE clients ADD COLUMN {IFNE} guest_token TEXT",
    "ALTER TABLE wpayments ADD COLUMN {IFNE} ws_label TEXT DEFAULT ''",
    "ALTER TABLE workspaces ADD COLUMN {IFNE} creds TEXT DEFAULT ''",
    "ALTER TABLE referrals ADD COLUMN {IFNE} reward_days INTEGER DEFAULT 0",
    "ALTER TABLE wpayments ALTER COLUMN workspace_id DROP NOT NULL",
)

PROTECTED_OWNER_EMAIL = 'nkheradia@gmail.com'   # master's own workspace — can never be deleted

def ws_protected(wid):
    r = db().execute("SELECT 1 FROM users WHERE workspace_id=? AND role='owner' AND email=? LIMIT 1",
                     (wid, PROTECTED_OWNER_EMAIL)).fetchone()
    return bool(r)

def _unit_days(unit):
    return {'days': 1, 'months': 30, 'years': 365}.get((unit or 'days').lower(), 1)

def _ws_creds(w, owner):
    c = {}
    try:
        if 'creds' in w.keys() and w['creds']:
            c = json.loads(w['creds'])
    except Exception:
        c = {}
    c.setdefault('email', owner['email'] if owner else '')
    c.setdefault('invite_code', w['invite_code'])
    c.setdefault('password', None)
    return c

def init_db():
    if USE_PG:
        conn = PgConn(DATABASE_URL)
        conn.executescript(SCHEMA.format(PK='BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY'))
        for mig in WS_MIGRATIONS:
            try: conn.executescript(mig.format(IFNE='IF NOT EXISTS'))
            except Exception: pass
        conn.close()
    else:
        conn = sqlite3.connect(DB)
        conn.executescript(SCHEMA.format(PK='INTEGER PRIMARY KEY'))
        for mig in WS_MIGRATIONS:
            try: conn.execute(mig.format(IFNE=''))
            except Exception: pass
        conn.commit()
        conn.close()

def get_ws_setting(wsid, key, default=None):
    r = db().execute("SELECT value FROM wsettings WHERE workspace_id=? AND key=?", (wsid, key)).fetchone()
    return r['value'] if r else default

def set_ws_setting(wsid, key, value):
    db().execute("""INSERT INTO wsettings(workspace_id,key,value) VALUES(?,?,?)
                    ON CONFLICT(workspace_id,key) DO UPDATE SET value=?""", (wsid, key, value, value))
    db().commit()

# ---------------- auth plumbing ----------------
def login_required(f):
    @wraps(f)
    def w(*a, **k):
        if 'uid' not in session: return jsonify(error="auth required"), 401
        return f(*a, **k)
    return w

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def w(*a, **k):
            if 'uid' not in session: return jsonify(error="auth required"), 401
            if session.get('role') not in roles: return jsonify(error="forbidden"), 403
            return f(*a, **k)
        return w
    return deco

def me():
    return db().execute("SELECT * FROM users WHERE id=?", (session['uid'],)).fetchone()

def wsid():
    return session['ws']

def ws_locked(ws_id):
    if ws_protected(ws_id): return None
    w = db().execute("SELECT wstatus, expires_on FROM workspaces WHERE id=?", (ws_id,)).fetchone()
    if not w: return "Workspace not found"
    if (w['wstatus'] or 'active') == 'suspended':
        return "This workspace has been suspended. Please contact BuzzFlow support."
    if w['expires_on'] and w['expires_on'] < today_ist().isoformat():
        return "Your subscription has ended. Please contact BuzzFlow to renew."
    return None

@app.before_request
def token_auth():
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        row = db().execute("""SELECT u.id, u.role, u.workspace_id FROM tokens k
                              JOIN users u ON u.id=k.user_id
                              WHERE k.token=? AND u.active=1""", (auth[7:],)).fetchone()
        if row:
            if ws_locked(row['workspace_id']):
                session.clear()
            else:
                session['uid'], session['role'], session['ws'] = row['id'], row['role'], row['workspace_id']

def issue_token(uid):
    token = secrets.token_hex(24)
    db().execute("INSERT INTO tokens(token,user_id,created_at) VALUES(?,?,?)",
                 (token, uid, now_ist().isoformat()))
    db().commit()
    return token

def userdict(u):
    return dict(id=u['id'], name=u['name'], role=u['role'], title=u['title'], email=u['email'])

def parse_join_token(tok):
    """'12.CODE' or a full URL containing '#join=12.CODE' -> (wsid, code) or None"""
    tok = (tok or '').strip()
    if '#join=' in tok:
        tok = tok.split('#join=')[-1]
    if '.' not in tok:
        return None
    a, b = tok.split('.', 1)
    if not a.isdigit() or not b:
        return None
    try:
        from urllib.parse import unquote
        return int(a), unquote(b)
    except Exception:
        return None

# ---------------- auth endpoints ----------------
@app.get('/api/join-info')
def join_info():
    p = parse_join_token(request.args.get('token', ''))
    if not p: return jsonify(error="Invalid invite link"), 400
    w = db().execute("SELECT id,name,invite_code FROM workspaces WHERE id=?", (p[0],)).fetchone()
    if not w or w['invite_code'] != p[1]:
        return jsonify(error="Invalid or expired invite link"), 404
    return jsonify(workspace_name=w['name'])

@app.post('/api/signup')
def signup():
    d = request.json or {}
    mode = d.get('mode') or 'create'
    name = (d.get('name') or '').strip()
    email = (d.get('email') or '').strip().lower()
    pw = d.get('password') or ''
    title = (d.get('title') or '').strip()
    if not name or not email or len(pw) < 4:
        return jsonify(error="Name, email and a password (min 4 chars) are required"), 400
    if '@' not in email or '.' not in email.split('@')[-1]:
        return jsonify(error="Please enter a valid email"), 400

    if mode == 'create':
        return jsonify(error="New workspaces are created by the BuzzFlow team. Contact us to get started!"), 403
    if True:  # join
        p = parse_join_token(d.get('join_token') or '')
        if not p:
            return jsonify(error="Paste the invite link or scan the QR from your owner"), 400
        w = db().execute("SELECT * FROM workspaces WHERE id=?", (p[0],)).fetchone()
        if not w or w['invite_code'] != p[1]:
            return jsonify(error="Invalid invite — ask your owner for a fresh link/QR"), 403
        role, ws = 'employee', w['id']
        title = title or 'Team Member'
        lock = ws_locked(ws)
        if lock:
            return jsonify(error=lock), 402
        wp = db().execute("SELECT plan FROM workspaces WHERE id=?", (ws,)).fetchone()
        pl = db().execute("SELECT max_members FROM plans WHERE name=?", (wp['plan'] or 'Pro',)).fetchone()
        if pl and pl['max_members']:
            cur_n = db().execute("SELECT COUNT(*) n FROM users WHERE workspace_id=? AND active=1", (ws,)).fetchone()['n']
            if cur_n >= pl['max_members']:
                return jsonify(error="This workspace has reached its member limit. Ask your owner to upgrade the plan."), 402

    salt = secrets.token_hex(8)
    try:
        cur = db().execute("""INSERT INTO users(workspace_id,name,email,salt,pw,role,title,created_at)
                              VALUES(?,?,?,?,?,?,?,?)""",
                           (ws, name, email, salt, hashpw(pw, salt), role, title, now_ist().isoformat()))
        db().commit()
    except Exception as ex:
        if not _is_unique_violation(ex):
            raise
        if mode == 'create':
            db().execute("DELETE FROM workspaces WHERE id=?", (ws,)); db().commit()
        return jsonify(error="This email is already registered — please sign in"), 400
    uid = cur.lastrowid
    session['uid'], session['role'], session['ws'] = uid, role, ws
    u = db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    return jsonify(token=issue_token(uid), user=userdict(u))

@app.post('/api/login')
def login():
    d = request.json or {}
    u = db().execute("SELECT * FROM users WHERE email=? AND active=1",
                     ((d.get('email') or '').strip().lower(),)).fetchone()
    if not u or hashpw(d.get('password',''), u['salt']) != u['pw']:
        return jsonify(error="Invalid email or password"), 401
    lock = ws_locked(u['workspace_id'])
    if lock:
        return jsonify(error=lock), 402
    session['uid'], session['role'], session['ws'] = u['id'], u['role'], u['workspace_id']
    return jsonify(token=issue_token(u['id']), user=userdict(u))

@app.post('/api/logout')
def logout():
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        db().execute("DELETE FROM tokens WHERE token=?", (auth[7:],)); db().commit()
    session.clear(); return jsonify(ok=True)

@app.get('/api/me')
@login_required
def whoami():
    w = db().execute("SELECT expires_on, plan FROM workspaces WHERE id=?", (wsid(),)).fetchone()
    warn = None
    if w and w['expires_on']:
        days = (date.fromisoformat(w['expires_on']) - today_ist()).days
        if 0 <= days <= 7:
            warn = f"Your {w['plan'] or ''} plan expires in {days} day{'s' if days != 1 else ''}. Please renew to avoid interruption."
    return jsonify(user=userdict(me()), expiry_warning=warn)

# ---------------- invite / QR ----------------
@app.get('/api/invite-code')
@role_required('owner')
def invite_get():
    from urllib.parse import quote
    w = db().execute("SELECT * FROM workspaces WHERE id=?", (wsid(),)).fetchone()
    token = f"{w['id']}.{quote(w['invite_code'])}"
    return jsonify(invite_code=w['invite_code'], workspace_id=w['id'], join_token=token,
                   workspace_name=w['name'])

@app.put('/api/invite-code')
@role_required('owner')
def invite_set():
    code = ((request.json or {}).get('invite_code') or '').strip()
    if len(code) < 4: return jsonify(error="Code must be at least 4 characters"), 400
    db().execute("UPDATE workspaces SET invite_code=? WHERE id=?", (code, wsid()))
    db().commit()
    return jsonify(ok=True)

# ---------------- company profile (per workspace) ----------------
@app.get('/api/company')
def company_get():
    if 'uid' not in session:
        return jsonify(company={})
    raw = get_ws_setting(wsid(), 'company')
    return jsonify(company=json.loads(raw) if raw else {})

@app.put('/api/company')
@role_required('owner')
def company_put():
    d = request.json or {}
    data = {k: (d.get(k) or '').strip() for k in
            ('name','tagline','phone','email','address','website','gst','report_phone')}
    logo = d.get('logo') or ''
    if logo and not logo.startswith('data:image/'):
        return jsonify(error="Invalid logo image"), 400
    if len(logo) > 900_000:
        return jsonify(error="Logo too large — please use a smaller image"), 400
    data['logo'] = logo
    set_ws_setting(wsid(), 'company', json.dumps(data))
    return jsonify(ok=True)

# ---------------- dashboard ----------------
@app.get('/api/dashboard')
@login_required
def dashboard():
    d = db(); role = session['role']; uid = session['uid']; ws = wsid()
    month = today_ist().strftime('%Y-%m')
    where_emp = "" if role in ('owner','admin') else f" AND t.assignee_id={uid}"
    counts = {r['status']: r['n'] for r in d.execute(
        f"SELECT status, COUNT(*) n FROM tasks t WHERE t.workspace_id=?{where_emp} GROUP BY status", (ws,))}
    overdue = d.execute(f"""SELECT COUNT(*) n FROM tasks t WHERE t.workspace_id=? AND t.status!='done'
                            AND t.due < ?{where_emp}""", (ws, today_ist().isoformat())).fetchone()['n']
    total = sum(counts.values()) or 1
    done = counts.get('done', 0)
    out = dict(counts=counts, overdue=overdue, completion=round(done*100/total),
               month=today_ist().strftime('%B %Y'))
    if role in ('owner','admin'):
        out['clients'] = d.execute("SELECT COUNT(*) n FROM clients WHERE workspace_id=? AND status='active'", (ws,)).fetchone()['n']
        out['team'] = d.execute("SELECT COUNT(*) n FROM users WHERE workspace_id=? AND active=1", (ws,)).fetchone()['n']
        if role == 'owner':
            out['revenue'] = d.execute("SELECT COALESCE(SUM(fee),0) s FROM clients WHERE workspace_id=? AND status='active'", (ws,)).fetchone()['s']
    rows = d.execute("""SELECT service, SUM(qty) n FROM tasks WHERE workspace_id=? AND status='done'
                        AND substr(done_at,1,7)=? GROUP BY service""", (ws, month)).fetchall()
    out['delivered'] = {r['service']: r['n'] for r in rows}
    targets = d.execute("""SELECT q.service, SUM(q.monthly_target) n FROM quotas q
                           JOIN clients c ON c.id=q.client_id
                           WHERE c.workspace_id=? AND c.status='active' GROUP BY q.service""", (ws,)).fetchall()
    out['targets'] = {r['service']: r['n'] for r in targets}
    return jsonify(out)

# ---------------- AI insights ----------------
def plural(s):
    return (s[:-1] + "ies") if s.endswith("y") else (s + "s")

@app.get('/api/insights')
@login_required
def insights():
    d = db(); ws = wsid(); today = today_ist(); month = today.strftime('%Y-%m')
    days_in_month = (date(today.year + (today.month==12), (today.month%12)+1, 1) - timedelta(days=1)).day
    pace = today.day / days_in_month
    tips = []
    rows = d.execute("""SELECT t.title, t.due, u.name a, c.name cl FROM tasks t
        LEFT JOIN users u ON u.id=t.assignee_id LEFT JOIN clients c ON c.id=t.client_id
        WHERE t.workspace_id=? AND t.status!='done' AND t.due < ? ORDER BY t.due LIMIT 5""",
        (ws, today.isoformat())).fetchall()
    for r in rows:
        tips.append(dict(level="red", icon="⚠️", text=f"Overdue: “{r['title']}” for {r['cl'] or '—'} — assigned to {r['a'] or 'nobody'}, was due {r['due']}."))
    for c in d.execute("SELECT * FROM clients WHERE workspace_id=? AND status='active'", (ws,)).fetchall():
        for q in d.execute("SELECT service, monthly_target FROM quotas WHERE client_id=?", (c['id'],)).fetchall():
            if q['monthly_target'] <= 0: continue
            done = d.execute("""SELECT COALESCE(SUM(qty),0) n FROM tasks WHERE client_id=? AND service=?
                AND status='done' AND substr(done_at,1,7)=?""", (c['id'], q['service'], month)).fetchone()['n']
            expected = q['monthly_target'] * pace
            if done >= q['monthly_target']:
                tips.append(dict(level="green", icon="✅", text=f"{c['name']}: {q['service']} quota complete ({done}/{q['monthly_target']}). Consider upsell."))
            elif done < expected * 0.6 and pace > 0.3:
                tips.append(dict(level="red", icon="🔥", text=f"{c['name']} at risk: only {done}/{q['monthly_target']} {plural(q['service'])} done, {round(pace*100)}% of month gone."))
            elif done < expected:
                tips.append(dict(level="amber", icon="⏳", text=f"{c['name']} slightly behind on {plural(q['service'])} ({done}/{q['monthly_target']})."))
    rows = d.execute("""SELECT u.name, COUNT(*) n FROM tasks t JOIN users u ON u.id=t.assignee_id
        WHERE t.workspace_id=? AND t.status IN('todo','doing','review') GROUP BY u.id ORDER BY n DESC""", (ws,)).fetchall()
    if len(rows) >= 2 and rows[0]['n'] >= rows[-1]['n'] * 2 and rows[0]['n'] >= 4:
        tips.append(dict(level="amber", icon="⚖️", text=f"Workload imbalance: {rows[0]['name']} has {rows[0]['n']} open tasks vs {rows[-1]['name']}'s {rows[-1]['n']}. Consider redistributing."))
    yest = (now_ist() - timedelta(days=1)).date().isoformat()
    quiet = d.execute("""SELECT name FROM users WHERE workspace_id=? AND role='employee' AND active=1 AND id NOT IN
        (SELECT DISTINCT user_id FROM updates WHERE workspace_id=? AND substr(created_at,1,10) >= ?)""",
        (ws, ws, yest)).fetchall()
    if quiet:
        tips.append(dict(level="amber", icon="💬", text="No update in 24h from: " + ", ".join(r['name'] for r in quiet) + "."))
    n = d.execute("SELECT COUNT(*) n FROM tasks WHERE workspace_id=? AND assignee_id IS NULL AND status!='done'", (ws,)).fetchone()['n']
    if n: tips.append(dict(level="red", icon="🎯", text=f"{n} task(s) have no assignee — assign them now."))
    pend = d.execute("""SELECT u.name, COUNT(*) n, COALESCE(SUM(t.qty),0) q FROM tasks t
        JOIN users u ON u.id=t.assignee_id WHERE t.workspace_id=? AND t.status!='done'
        GROUP BY u.id ORDER BY n DESC LIMIT 3""", (ws,)).fetchall()
    for p in pend:
        tips.append(dict(level="amber", icon="🏃", text=f"{p['name']} has {p['n']} pending task(s) covering {p['q']} deliverable(s) — cheer them on!"))
    if not tips:
        tips.append(dict(level="green", icon="🌿", text="All clear! Add clients and tasks to get insights."))
    order = {"red":0, "amber":1, "green":2}
    tips.sort(key=lambda t: order[t['level']])
    return jsonify(insights=tips[:12])

# ---------------- WhatsApp daily report ----------------
@app.get('/api/report/daily')
@role_required('owner','admin')
def daily_report():
    d = db(); ws = wsid()
    raw = get_ws_setting(ws, 'company')
    company = json.loads(raw) if raw else {}
    phone = ''.join(ch for ch in (company.get('report_phone') or '') if ch.isdigit())
    if not phone:
        return jsonify(error="Set the WhatsApp report number first (Settings → Company profile)"), 400
    if len(phone) == 10:
        phone = '91' + phone
    today = today_ist(); tstr = today.isoformat(); month = today.strftime('%Y-%m')
    nice = today.strftime('%d %b %Y')
    co = company.get('name') or 'BuzzFlow'
    L = [f"📊 *{co} Daily Report — {nice}*", ""]
    done_today = d.execute("""SELECT t.title,t.qty,t.service,u.name a,c.name cl FROM tasks t
        LEFT JOIN users u ON u.id=t.assignee_id LEFT JOIN clients c ON c.id=t.client_id
        WHERE t.workspace_id=? AND t.status='done' AND substr(t.done_at,1,10)=?""", (ws, tstr)).fetchall()
    total_q = sum(r['qty'] for r in done_today)
    L.append(f"✅ *Completed today:* {len(done_today)} task(s), {total_q} deliverable(s)")
    overdue = d.execute("SELECT COUNT(*) n FROM tasks WHERE workspace_id=? AND status!='done' AND due<?",
                        (ws, tstr)).fetchone()['n']
    L.append(f"🔥 *Overdue:* {overdue}")
    quotas = d.execute("""SELECT q.service, SUM(q.monthly_target) tgt FROM quotas q
        JOIN clients c ON c.id=q.client_id WHERE c.workspace_id=? AND c.status='active'
        GROUP BY q.service""", (ws,)).fetchall()
    if quotas:
        parts = []
        for q in quotas:
            done = d.execute("""SELECT COALESCE(SUM(qty),0) n FROM tasks WHERE workspace_id=? AND service=?
                AND status='done' AND substr(done_at,1,7)=?""", (ws, q['service'], month)).fetchone()['n']
            parts.append(f"{plural(q['service'])} {done}/{q['tgt']}")
        L.append("📦 *Month:* " + " · ".join(parts))
    L.append("")
    L.append("👥 *EMPLOYEE-WISE:*")
    emps = d.execute("SELECT id,name FROM users WHERE workspace_id=? AND active=1 AND role!='owner' ORDER BY name", (ws,)).fetchall()
    if not emps:
        L.append("• No team members yet.")
    for e in emps:
        done_rows = d.execute("""SELECT t.qty,t.service,c.name cl FROM tasks t
            LEFT JOIN clients c ON c.id=t.client_id
            WHERE t.assignee_id=? AND t.status='done' AND substr(t.done_at,1,10)=?""", (e['id'], tstr)).fetchall()
        pending = d.execute("SELECT COUNT(*) n FROM tasks WHERE assignee_id=? AND status!='done'", (e['id'],)).fetchone()['n']
        if done_rows:
            done_txt = ", ".join(f"{r['qty']} {plural(r['service']) if r['qty']>1 else r['service']}" +
                                 (f" ({r['cl']})" if r['cl'] else "") for r in done_rows)
            L.append(f"• *{e['name']}* — ✅ {done_txt} completed · {pending} pending")
        else:
            L.append(f"• *{e['name']}* — ❌ nothing completed today · {pending} pending")
    L.append("")
    posted = d.execute("""SELECT DISTINCT u.name FROM updates up JOIN users u ON u.id=up.user_id
        WHERE up.workspace_id=? AND substr(up.created_at,1,10)=?""", (ws, tstr)).fetchall()
    silent = d.execute("""SELECT name FROM users WHERE workspace_id=? AND active=1 AND role='employee' AND id NOT IN
        (SELECT user_id FROM updates WHERE workspace_id=? AND substr(created_at,1,10)=?)""", (ws, ws, tstr)).fetchall()
    L.append("💬 *Updates posted:* " + (", ".join(r['name'] for r in posted) or "none"))
    if silent:
        L.append("🔕 *Silent today:* " + ", ".join(r['name'] for r in silent))
    L.append("")
    L.append("_Sent via BuzzFlow — Agency OS_")
    return jsonify(text="\n".join(L), phone=phone)

# ---------------- calendar ----------------
@app.get('/api/calendar')
@login_required
def calendar_data():
    d = db(); ws = wsid()
    month = request.args.get('month') or today_ist().strftime('%Y-%m')
    due = d.execute("""SELECT t.id,t.title,t.status,t.qty,t.service,t.due AS "day",u.name assignee,c.name client
        FROM tasks t LEFT JOIN users u ON u.id=t.assignee_id LEFT JOIN clients c ON c.id=t.client_id
        WHERE t.workspace_id=? AND substr(t.due,1,7)=?""", (ws, month)).fetchall()
    done = d.execute("""SELECT t.id,t.title,t.status,t.qty,t.service,substr(t.done_at,1,10) AS "day",u.name assignee,c.name client
        FROM tasks t LEFT JOIN users u ON u.id=t.assignee_id LEFT JOIN clients c ON c.id=t.client_id
        WHERE t.workspace_id=? AND t.done_at IS NOT NULL AND substr(t.done_at,1,7)=?""", (ws, month)).fetchall()
    ups = d.execute("""SELECT up.id, substr(up.created_at,1,10) AS "day", up.text, u.name, c.name client
        FROM updates up JOIN users u ON u.id=up.user_id LEFT JOIN clients c ON c.id=up.client_id
        WHERE up.workspace_id=? AND substr(up.created_at,1,7)=?""", (ws, month)).fetchall()
    return jsonify(month=month, due=[dict(r) for r in due], done=[dict(r) for r in done],
                   updates=[dict(r) for r in ups], today=today_ist().isoformat())

# ---------------- clients ----------------
def clean_service(s):
    s = (s or 'Other').strip()[:40]
    return s if s else 'Other'

@app.get('/api/clients')
@login_required
def clients_list():
    d = db(); ws = wsid(); month = today_ist().strftime('%Y-%m')
    rows = d.execute("SELECT * FROM clients WHERE workspace_id=? ORDER BY status='active' DESC, name", (ws,)).fetchall()
    out = []
    for c in rows:
        item = dict(id=c['id'], name=c['name'], package=c['package'], status=c['status'],
                    contact_name=c['contact_name'], contact_phone=c['contact_phone'], notes=c['notes'])
        if session['role'] == 'owner': item['fee'] = c['fee']
        qs = d.execute("SELECT service, monthly_target FROM quotas WHERE client_id=?", (c['id'],)).fetchall()
        prog = []
        for q in qs:
            done = d.execute("""SELECT COALESCE(SUM(qty),0) n FROM tasks WHERE client_id=? AND service=?
                AND status='done' AND substr(done_at,1,7)=?""", (c['id'], q['service'], month)).fetchone()['n']
            prog.append(dict(service=q['service'], target=q['monthly_target'], done=done))
        item['quotas'] = prog
        out.append(item)
    return jsonify(clients=out)

@app.post('/api/clients')
@role_required('owner','admin')
def clients_add():
    d = request.json or {}
    if not (d.get('name') or '').strip(): return jsonify(error="name required"), 400
    cur = db().execute("""INSERT INTO clients(workspace_id,name,package,fee,contact_name,contact_phone,notes,created_at)
                          VALUES(?,?,?,?,?,?,?,?)""",
        (wsid(), d['name'].strip(), d.get('package',''), int(d.get('fee') or 0), d.get('contact_name',''),
         d.get('contact_phone',''), d.get('notes',''), now_ist().isoformat()))
    cid = cur.lastrowid
    for q in d.get('quotas', []):
        if int(q.get('target') or 0) > 0:
            svc = clean_service(q.get('service'))
            learn_service(svc)
            db().execute("INSERT INTO quotas(client_id,service,monthly_target) VALUES(?,?,?)",
                         (cid, svc, int(q['target'])))
    db().commit()
    return jsonify(id=cid)

def own_client(cid):
    return db().execute("SELECT * FROM clients WHERE id=? AND workspace_id=?", (cid, wsid())).fetchone()

@app.put('/api/clients/<int:cid>')
@role_required('owner','admin')
def clients_edit(cid):
    if not own_client(cid): return jsonify(error="not found"), 404
    d = request.json or {}
    db().execute("""UPDATE clients SET name=?,package=?,fee=?,contact_name=?,contact_phone=?,notes=?,status=?
                    WHERE id=? AND workspace_id=?""",
        (d.get('name'), d.get('package',''), int(d.get('fee') or 0), d.get('contact_name',''),
         d.get('contact_phone',''), d.get('notes',''), d.get('status','active'), cid, wsid()))
    if 'quotas' in d:
        db().execute("DELETE FROM quotas WHERE client_id=?", (cid,))
        for q in d['quotas']:
            if int(q.get('target') or 0) > 0:
                db().execute("INSERT INTO quotas(client_id,service,monthly_target) VALUES(?,?,?)",
                             (cid, clean_service(q.get('service')), int(q['target'])))
    db().commit()
    return jsonify(ok=True)

@app.post('/api/clients/<int:cid>/quotas')
@role_required('owner','admin')
def quota_upsert(cid):
    if not own_client(cid): return jsonify(error="not found"), 404
    d = request.json or {}
    svc = clean_service(d.get('service'))
    target = int(d.get('target') or 0)
    if target <= 0: return jsonify(error="Enter a monthly target"), 400
    learn_service(svc)
    ex = db().execute("SELECT id FROM quotas WHERE client_id=? AND LOWER(service)=LOWER(?)", (cid, svc)).fetchone()
    if ex:
        db().execute("UPDATE quotas SET monthly_target=? WHERE id=?", (target, ex['id']))
    else:
        db().execute("INSERT INTO quotas(client_id,service,monthly_target) VALUES(?,?,?)", (cid, svc, target))
    db().commit()
    return jsonify(ok=True)

@app.delete('/api/clients/<int:cid>')
@role_required('owner')
def clients_del(cid):
    if not own_client(cid): return jsonify(error="not found"), 404
    db().execute("DELETE FROM clients WHERE id=? AND workspace_id=?", (cid, wsid())); db().commit()
    return jsonify(ok=True)

# ---------------- tasks ----------------
@app.get('/api/tasks')
@login_required
def tasks_list():
    d = db()
    q = """SELECT t.*, u.name assignee, c.name client FROM tasks t
           LEFT JOIN users u ON u.id=t.assignee_id LEFT JOIN clients c ON c.id=t.client_id
           WHERE t.workspace_id=?"""
    args = [wsid()]
    if request.args.get('mine') == '1':
        q += " AND t.assignee_id=?"; args.append(session['uid'])
    q += " ORDER BY CASE WHEN t.due IS NULL OR t.due='' THEN 1 ELSE 0 END, t.due, CASE t.priority WHEN 'urgent' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END"
    rows = [dict(r) for r in d.execute(q, args).fetchall()]
    return jsonify(tasks=rows, today=today_ist().isoformat())

@app.post('/api/tasks')
@login_required
def tasks_add():
    d = request.json or {}
    if not (d.get('title') or '').strip(): return jsonify(error="title required"), 400
    if session['role'] == 'employee':
        d['assignee_id'] = session['uid']
    cid = d.get('client_id')
    if cid and not own_client(cid): return jsonify(error="client not found"), 404
    aid = d.get('assignee_id')
    if aid:
        au = db().execute("SELECT 1 FROM users WHERE id=? AND workspace_id=?", (aid, wsid())).fetchone()
        if not au: return jsonify(error="assignee not found"), 404
    status = d.get('status','todo')
    done_at = now_ist().isoformat() if status == 'done' else None
    learn_service(clean_service(d.get('service')))
    cur = db().execute("""INSERT INTO tasks(workspace_id,title,client_id,service,qty,assignee_id,status,priority,due,notes,created_by,created_at,done_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (wsid(), d['title'].strip(), cid, clean_service(d.get('service')), int(d.get('qty') or 1),
         aid, status, d.get('priority','medium'), d.get('due'),
         d.get('notes',''), session['uid'], now_ist().isoformat(), done_at))
    db().commit()
    return jsonify(id=cur.lastrowid)

@app.put('/api/tasks/<int:tid>')
@login_required
def tasks_edit(tid):
    d = request.json or {}; conn = db()
    t = conn.execute("SELECT * FROM tasks WHERE id=? AND workspace_id=?", (tid, wsid())).fetchone()
    if not t: return jsonify(error="not found"), 404
    if session['role'] == 'employee' and t['assignee_id'] != session['uid']:
        return jsonify(error="forbidden"), 403
    fields = {}
    allowed = ['title','client_id','service','qty','status','priority','due','notes'] + \
              (['assignee_id'] if session['role'] in ('owner','admin') else [])
    for k in allowed:
        if k in d: fields[k] = d[k]
    if 'service' in fields:
        fields['service'] = clean_service(fields['service'])
        learn_service(fields['service'])
    if fields.get('client_id') and not own_client(fields['client_id']):
        return jsonify(error="client not found"), 404
    if fields.get('assignee_id'):
        au = conn.execute("SELECT 1 FROM users WHERE id=? AND workspace_id=?", (fields['assignee_id'], wsid())).fetchone()
        if not au: return jsonify(error="assignee not found"), 404
    if fields.get('status') == 'done' and t['status'] != 'done':
        fields['done_at'] = now_ist().isoformat()
    if fields.get('status') and fields['status'] != 'done' and t['status'] == 'done':
        fields['done_at'] = None
    if fields:
        sets = ",".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE tasks SET {sets} WHERE id=?", (*fields.values(), tid))
        conn.commit()
    return jsonify(ok=True)

@app.delete('/api/tasks/<int:tid>')
@role_required('owner','admin')
def tasks_del(tid):
    t = db().execute("SELECT 1 FROM tasks WHERE id=? AND workspace_id=?", (tid, wsid())).fetchone()
    if not t: return jsonify(error="not found"), 404
    db().execute("DELETE FROM tasks WHERE id=?", (tid,)); db().commit()
    return jsonify(ok=True)

# ---------------- updates ----------------
@app.get('/api/updates')
@login_required
def updates_list():
    rows = db().execute("""SELECT up.*, u.name, u.title, c.name client FROM updates up
        JOIN users u ON u.id=up.user_id LEFT JOIN clients c ON c.id=up.client_id
        WHERE up.workspace_id=? ORDER BY up.created_at DESC LIMIT 60""", (wsid(),)).fetchall()
    return jsonify(updates=[dict(r) for r in rows])

@app.post('/api/updates')
@login_required
def updates_add():
    d = request.json or {}
    if not (d.get('text') or '').strip(): return jsonify(error="text required"), 400
    cid = d.get('client_id')
    if cid and not own_client(cid): return jsonify(error="client not found"), 404
    db().execute("INSERT INTO updates(workspace_id,user_id,client_id,text,created_at) VALUES(?,?,?,?,?)",
                 (wsid(), session['uid'], cid, d['text'].strip(), now_ist().isoformat()))
    db().commit()
    return jsonify(ok=True)

# ---------------- team ----------------
@app.get('/api/team')
@login_required
def team_list():
    d = db(); ws = wsid()
    rows = d.execute("""SELECT id,name,email,role,title,active FROM users WHERE workspace_id=?
                        ORDER BY role='owner' DESC, role='admin' DESC, name""", (ws,)).fetchall()
    out = []
    for u in rows:
        open_n = d.execute("SELECT COUNT(*) n FROM tasks WHERE assignee_id=? AND status!='done'", (u['id'],)).fetchone()['n']
        done_m = d.execute("SELECT COALESCE(SUM(qty),0) n FROM tasks WHERE assignee_id=? AND status='done' AND substr(done_at,1,7)=?",
                           (u['id'], today_ist().strftime('%Y-%m'))).fetchone()['n']
        out.append(dict(id=u['id'], name=u['name'], email=u['email'], role=u['role'], title=u['title'],
                        active=u['active'], open_tasks=open_n, delivered_month=done_m))
    return jsonify(team=out)

@app.post('/api/team')
@role_required('owner','admin')
def team_add():
    d = request.json or {}
    if not all((d.get(k) or '').strip() for k in ('name','email','password')):
        return jsonify(error="name, email, password required"), 400
    w = db().execute("SELECT plan FROM workspaces WHERE id=?", (wsid(),)).fetchone()
    p = db().execute("SELECT max_members FROM plans WHERE name=?", (w['plan'] or 'Pro',)).fetchone()
    if p and p['max_members']:
        cur_n = db().execute("SELECT COUNT(*) n FROM users WHERE workspace_id=? AND active=1", (wsid(),)).fetchone()['n']
        if cur_n >= p['max_members']:
            return jsonify(error=f"Your {w['plan']} plan allows up to {p['max_members']} members. Contact BuzzFlow to upgrade."), 402
    salt = secrets.token_hex(8)
    try:
        db().execute("""INSERT INTO users(workspace_id,name,email,salt,pw,role,title,created_at)
                        VALUES(?,?,?,?,?,?,?,?)""",
            (wsid(), d['name'].strip(), d['email'].strip().lower(), salt, hashpw(d['password'], salt),
             'employee', d.get('title',''), now_ist().isoformat()))
        db().commit()
    except Exception as ex:
        if not _is_unique_violation(ex):
            raise
        return jsonify(error="email already exists"), 400
    return jsonify(ok=True)

def own_user(uid):
    return db().execute("SELECT * FROM users WHERE id=? AND workspace_id=?", (uid, wsid())).fetchone()

@app.put('/api/team/<int:uid>')
@role_required('owner','admin')
def team_edit(uid):
    d = request.json or {}; conn = db()
    u = own_user(uid)
    if not u: return jsonify(error="not found"), 404
    if u['role'] == 'owner' and session['role'] != 'owner': return jsonify(error="forbidden"), 403
    conn.execute("UPDATE users SET name=?, title=?, active=? WHERE id=?",
                 (d.get('name', u['name']), d.get('title', u['title']), int(d.get('active', u['active'])), uid))
    if d.get('password'):
        salt = secrets.token_hex(8)
        conn.execute("UPDATE users SET salt=?, pw=? WHERE id=?", (salt, hashpw(d['password'], salt), uid))
    conn.commit()
    return jsonify(ok=True)

@app.put('/api/team/<int:uid>/role')
@role_required('owner')
def team_role(uid):
    new_role = ((request.json or {}).get('role') or '').strip()
    if new_role not in ('admin', 'employee'):
        return jsonify(error="Role must be admin or employee"), 400
    u = own_user(uid)
    if not u: return jsonify(error="not found"), 404
    if u['role'] == 'owner': return jsonify(error="the owner role cannot be changed"), 403
    db().execute("UPDATE users SET role=? WHERE id=?", (new_role, uid))
    db().execute("DELETE FROM tokens WHERE user_id=?", (uid,))
    db().commit()
    return jsonify(ok=True)

@app.delete('/api/team/<int:uid>')
@role_required('owner','admin')
def team_del(uid):
    conn = db()
    u = own_user(uid)
    if not u: return jsonify(error="not found"), 404
    if uid == session['uid']: return jsonify(error="you cannot delete yourself"), 400
    if u['role'] == 'owner': return jsonify(error="the owner account cannot be deleted"), 403
    if u['role'] == 'admin' and session['role'] != 'owner':
        return jsonify(error="only the owner can delete a manager"), 403
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    return jsonify(ok=True)

# ---------------- work report (data + excel) ----------------
def build_report_data():
    d = db(); ws = wsid()
    month = today_ist().strftime('%Y-%m')
    month_name = today_ist().strftime('%B %Y')
    # per-employee summary
    emps = d.execute("""SELECT id,name,title,role FROM users WHERE workspace_id=? AND active=1
                        ORDER BY role='owner' DESC, name""", (ws,)).fetchall()
    emp_rows = []
    for e in emps:
        done_m = d.execute("""SELECT COALESCE(SUM(qty),0) n FROM tasks WHERE assignee_id=? AND status='done'
                              AND substr(done_at,1,7)=?""", (e['id'], month)).fetchone()['n']
        open_n = d.execute("SELECT COUNT(*) n FROM tasks WHERE assignee_id=? AND status!='done'", (e['id'],)).fetchone()['n']
        overdue_n = d.execute("""SELECT COUNT(*) n FROM tasks WHERE assignee_id=? AND status!='done'
                                 AND due < ?""", (e['id'], today_ist().isoformat())).fetchone()['n']
        svc = d.execute("""SELECT service, SUM(qty) n FROM tasks WHERE assignee_id=? AND status='done'
                           AND substr(done_at,1,7)=? GROUP BY service ORDER BY n DESC""", (e['id'], month)).fetchall()
        emp_rows.append(dict(name=e['name'], title=e['title'] or '', role=e['role'],
                             delivered=done_m, open=open_n, overdue=overdue_n,
                             services=', '.join(f"{r['n']} {r['service']}" for r in svc) or '—'))
    # per-client progress — ALL active clients, even without quotas
    cli_rows = []
    for c in d.execute("SELECT * FROM clients WHERE workspace_id=? AND status='active' ORDER BY name", (ws,)).fetchall():
        quotas = d.execute("SELECT service, monthly_target FROM quotas WHERE client_id=?", (c['id'],)).fetchall()
        seen = set()
        for q in quotas:
            done = d.execute("""SELECT COALESCE(SUM(qty),0) n FROM tasks WHERE client_id=? AND service=?
                AND status='done' AND substr(done_at,1,7)=?""", (c['id'], q['service'], month)).fetchone()['n']
            seen.add(q['service'].lower())
            cli_rows.append(dict(client=c['name'], service=q['service'],
                                 done=done, target=q['monthly_target'],
                                 pct=round(done*100/q['monthly_target']) if q['monthly_target'] else 0))
        # services delivered this month that have no quota row
        extra = d.execute("""SELECT service, SUM(qty) n FROM tasks WHERE client_id=? AND status='done'
            AND substr(done_at,1,7)=? GROUP BY service""", (c['id'], month)).fetchall()
        for e in extra:
            if e['service'].lower() not in seen:
                cli_rows.append(dict(client=c['name'], service=e['service'],
                                     done=e['n'], target=None, pct=None))
        if not quotas and not extra:
            cli_rows.append(dict(client=c['name'], service='— no work recorded —',
                                 done=None, target=None, pct=None))
    # all tasks detail
    task_rows = [dict(r) for r in d.execute("""SELECT t.title, t.service, t.qty, t.status, t.priority, t.due,
        substr(t.done_at,1,10) done_on, u.name assignee, c.name client
        FROM tasks t LEFT JOIN users u ON u.id=t.assignee_id LEFT JOIN clients c ON c.id=t.client_id
        WHERE t.workspace_id=? ORDER BY CASE WHEN t.due IS NULL THEN 1 ELSE 0 END, t.due""", (ws,)).fetchall()]
    totals = dict(
        total_tasks=len(task_rows),
        done=len([t for t in task_rows if t['status']=='done']),
        open=len([t for t in task_rows if t['status']!='done']),
        overdue=len([t for t in task_rows if t['status']!='done' and t['due'] and t['due'] < today_ist().isoformat()]),
        delivered_month=sum(e['delivered'] for e in emp_rows))
    raw = get_ws_setting(ws, 'company')
    co = json.loads(raw) if raw else {}
    return dict(month=month_name, company=co.get('name') or 'BuzzFlow',
                generated=now_ist().strftime('%d %b %Y, %I:%M %p IST'),
                totals=totals, employees=emp_rows, clients=cli_rows, tasks=task_rows)

@app.get('/api/report/work')
@role_required('owner','admin')
def work_report_data():
    return jsonify(build_report_data())

@app.get('/api/report/work.xlsx')
@role_required('owner','admin')
def work_report_xlsx():
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    r = build_report_data()
    wb = Workbook()
    orange_fill = PatternFill('solid', fgColor='FF5A1F')
    dark_fill = PatternFill('solid', fgColor='3C4048')
    zebra_fill = PatternFill('solid', fgColor='FAF8F4')
    white_bold = Font(color='FFFFFF', bold=True)
    red_font = Font(color='D64545', bold=True)
    green_font = Font(color='2E8B57', bold=True)
    orange_font = Font(color='FF5A1F', bold=True)
    mut_font = Font(color='787D87', italic=True)
    thin = Border(bottom=Side(style='thin', color='DEDAD2'))

    def style_header(ws_, row, ncols, fill, font):
        for c in range(1, ncols+1):
            cell = ws_.cell(row=row, column=c)
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal='left')

    def autow(ws_):
        for col in ws_.columns:
            width = max(len(str(c.value or '')) for c in col) + 3
            ws_.column_dimensions[col[0].column_letter].width = min(width, 46)

    def plural_svc(n, svc):
        base = svc[:-1] if svc.lower().endswith('s') else svc
        return f"{n} {base}" if n == 1 else f"{n} {base}s"

    # ---- Sheet 1: Summary ----
    s1 = wb.active; s1.title = 'Summary'
    s1.append([r['company'] + ' — Monthly Work Report'])
    s1['A1'].font = Font(bold=True, size=14, color='FF5A1F')
    s1.append([r['month'] + '  ·  Generated ' + r['generated']])
    s1['A2'].font = mut_font
    s1.append([])
    s1.append(['Total tasks', 'Completed', 'Still open', 'Overdue', 'Delivered this month'])
    style_header(s1, 4, 5, dark_fill, white_bold)
    T = r['totals']
    s1.append([T['total_tasks'], T['done'], T['open'], T['overdue'], T['delivered_month']])
    s1.cell(row=5, column=2).font = green_font
    s1.cell(row=5, column=4).font = red_font if T['overdue'] else green_font
    s1.cell(row=5, column=5).font = orange_font
    autow(s1)

    # ---- Sheet 2: Monthly Work (client wise, pending/extra) ----
    s2 = wb.create_sheet('Monthly Work')
    s2.append(['Client', 'Service', 'Done', 'Target', 'Progress %', 'Pending', 'Extra delivered'])
    style_header(s2, 1, 7, orange_fill, white_bold)
    byc = {}
    for c in r['clients']:
        byc.setdefault(c['client'], []).append(c)
    ri = 2
    for cn, rows in byc.items():
        for qi, q in enumerate(rows):
            pend_txt, extra_txt, pct_txt = '', '', ''
            done_v = q['done'] if q['done'] is not None else ''
            tgt_v = q['target'] if q['target'] else '—'
            if q['target']:
                pct_txt = f"{q['pct']}%"
                if q['done'] < q['target']:
                    pend_txt = plural_svc(q['target'] - q['done'], q['service'])
                elif q['done'] > q['target']:
                    extra_txt = '+' + plural_svc(q['done'] - q['target'], q['service'])
            elif q['done'] is not None:
                pct_txt = 'no target set'
                extra_txt = '+' + plural_svc(q['done'], q['service'])
            else:
                pct_txt = 'no work recorded'
            s2.append([cn if qi == 0 else '', q['service'], done_v, tgt_v, pct_txt, pend_txt, extra_txt])
            if qi == 0:
                s2.cell(row=ri, column=1).font = orange_font
            if pend_txt: s2.cell(row=ri, column=6).font = red_font
            if extra_txt: s2.cell(row=ri, column=7).font = green_font
            if pct_txt.endswith('%') and q['done'] >= (q['target'] or 0):
                s2.cell(row=ri, column=5).font = green_font
            if pct_txt in ('no target set', 'no work recorded'):
                s2.cell(row=ri, column=5).font = mut_font
            if ri % 2 == 0:
                for c2 in range(1, 8): s2.cell(row=ri, column=c2).fill = zebra_fill
            for c2 in range(1, 8): s2.cell(row=ri, column=c2).border = thin
            ri += 1
    autow(s2)

    # ---- Sheet 3: Pending Work ----
    s3 = wb.create_sheet('Pending Work')
    s3.append(['Task', 'Client', 'Assigned to', 'Status', 'Due date', 'Qty', 'Overdue?'])
    style_header(s3, 1, 7, orange_fill, white_bold)
    tstr = today_ist().isoformat()
    ri = 2
    for t in [t for t in r['tasks'] if t['status'] != 'done']:
        over = bool(t['due'] and t['due'] < tstr)
        s3.append([t['title'] + (f" ×{t['qty']}" if t['qty'] > 1 else ''), t['client'] or '—',
                   t['assignee'] or 'UNASSIGNED', t['status'], t['due'] or '—', t['qty'],
                   'YES' if over else ''])
        if not t['assignee']: s3.cell(row=ri, column=3).font = red_font
        if over:
            s3.cell(row=ri, column=5).font = red_font
            s3.cell(row=ri, column=7).font = red_font
        if ri % 2 == 0:
            for c2 in range(1, 8): s3.cell(row=ri, column=c2).fill = zebra_fill
        ri += 1
    if ri == 2:
        s3.append(['All caught up — no pending tasks!'])
        s3.cell(row=2, column=1).font = green_font
    autow(s3)

    # ---- Sheet 4: Employee Performance ----
    s4 = wb.create_sheet('Employee Performance')
    s4.append(['Employee', 'Role', 'Delivered', 'Open', 'Overdue', 'Share of month %', 'Work breakdown'])
    style_header(s4, 1, 7, orange_fill, white_bold)
    tot = max(1, T['delivered_month'])
    ri = 2
    for e in r['employees']:
        share = round(e['delivered'] * 100 / tot)
        s4.append([e['name'], e['title'] or e['role'], e['delivered'], e['open'], e['overdue'],
                   f"{share}%", e['services']])
        s4.cell(row=ri, column=1).font = Font(bold=True)
        s4.cell(row=ri, column=3).font = green_font
        if e['overdue']: s4.cell(row=ri, column=5).font = red_font
        if share >= 40: s4.cell(row=ri, column=6).font = green_font
        elif share > 0: s4.cell(row=ri, column=6).font = orange_font
        if ri % 2 == 0:
            for c2 in range(1, 8): s4.cell(row=ri, column=c2).fill = zebra_fill
        ri += 1
    autow(s4)

    buf = io.BytesIO()
    wb.save(buf)
    return jsonify(filename=f"BuzzFlow-Report-{today_ist().isoformat()}.xlsx",
                   b64=base64.b64encode(buf.getvalue()).decode())

# ---------------- services (per-workspace list) ----------------
def ws_services():
    if not get_ws_setting(wsid(), 'svc_seeded'):
        for n in SERVICE_TYPES:
            ex = db().execute("SELECT 1 FROM services WHERE workspace_id=? AND LOWER(name)=LOWER(?)",
                              (wsid(), n)).fetchone()
            if not ex:
                db().execute("INSERT INTO services(workspace_id,name) VALUES(?,?)", (wsid(), n))
        db().commit()
        set_ws_setting(wsid(), 'svc_seeded', '1')
    return db().execute("SELECT id,name FROM services WHERE workspace_id=? ORDER BY name", (wsid(),)).fetchall()

def learn_service(name):
    """Auto-add manually typed services to the workspace list."""
    name = (name or '').strip()[:40]
    if not name or name == 'Other':
        return
    ex = db().execute("SELECT 1 FROM services WHERE workspace_id=? AND LOWER(name)=LOWER(?)",
                      (wsid(), name)).fetchone()
    if not ex:
        db().execute("INSERT INTO services(workspace_id,name) VALUES(?,?)", (wsid(), name))
        db().commit()

@app.get('/api/services')
@login_required
def services_list():
    return jsonify(services=[dict(r) for r in ws_services()])

@app.post('/api/services')
@role_required('owner','admin')
def services_add():
    name = ((request.json or {}).get('name') or '').strip()[:40]
    if not name: return jsonify(error="name required"), 400
    ex = db().execute("SELECT 1 FROM services WHERE workspace_id=? AND LOWER(name)=LOWER(?)", (wsid(), name)).fetchone()
    if ex: return jsonify(error="Service already exists"), 400
    cur = db().execute("INSERT INTO services(workspace_id,name) VALUES(?,?)", (wsid(), name))
    db().commit()
    return jsonify(id=cur.lastrowid)

@app.delete('/api/services/<int:sid>')
@role_required('owner','admin')
def services_del(sid):
    r = db().execute("SELECT 1 FROM services WHERE id=? AND workspace_id=?", (sid, wsid())).fetchone()
    if not r: return jsonify(error="not found"), 404
    db().execute("DELETE FROM services WHERE id=?", (sid,)); db().commit()
    return jsonify(ok=True)

# ================= MASTER PANEL =================
MASTER_PASSWORD = os.environ.get('MASTER_PASSWORD', 'change-me-please')

def master_required(f):
    @wraps(f)
    def w(*a, **k):
        auth = request.headers.get('X-Master-Token', '')
        row = db().execute("SELECT 1 FROM msessions WHERE token=?", (auth,)).fetchone() if auth else None
        if not row: return jsonify(error="master auth required"), 401
        return f(*a, **k)
    return w

@app.post('/api/master/login')
def master_login():
    import time as _t
    _t.sleep(1.2)
    if (request.json or {}).get('password') != MASTER_PASSWORD:
        return jsonify(error="Wrong master password"), 401
    token = secrets.token_hex(24)
    db().execute("INSERT INTO msessions(token,created_at) VALUES(?,?)", (token, now_ist().isoformat()))
    db().commit()
    return jsonify(token=token)

def seed_plans():
    n = db().execute("SELECT COUNT(*) n FROM plans").fetchone()['n']
    if not n:
        for p in (('Trial', 0, 0, 5), ('Starter', 999, 9990, 5),
                  ('Growth', 1999, 19990, 15), ('Pro', 3499, 34990, 0)):
            db().execute("INSERT INTO plans(name,price_m,price_y,max_members) VALUES(?,?,?,?)", p)
        db().commit()

@app.get('/api/master/overview')
@master_required
def master_overview():
    seed_plans()
    d = db(); today = today_ist().isoformat()
    wss = d.execute("SELECT * FROM workspaces ORDER BY id").fetchall()
    out = []
    month = today_ist().strftime('%Y-%m')
    for w in wss:
        owner = d.execute("SELECT name,email FROM users WHERE workspace_id=? AND role='owner' LIMIT 1", (w['id'],)).fetchone()
        if ws_protected(w['id']) and w['expires_on']:
            d.execute("UPDATE workspaces SET expires_on=NULL WHERE id=?", (w['id'],)); d.commit()
            w = d.execute("SELECT * FROM workspaces WHERE id=?", (w['id'],)).fetchone()
        members = d.execute("SELECT COUNT(*) n FROM users WHERE workspace_id=? AND active=1", (w['id'],)).fetchone()['n']
        clients = d.execute("SELECT COUNT(*) n FROM clients WHERE workspace_id=?", (w['id'],)).fetchone()['n']
        tasks_m = d.execute("SELECT COUNT(*) n FROM tasks WHERE workspace_id=? AND substr(created_at,1,7)=?", (w['id'], month)).fetchone()['n']
        exp = w['expires_on']
        days_left = (date.fromisoformat(exp) - today_ist()).days if exp else None
        if (w['wstatus'] or 'active') == 'suspended': status = 'suspended'
        elif exp and exp < today: status = 'expired'
        elif days_left is not None and days_left <= 15: status = 'expiring'
        elif (w['plan'] or '') == 'Trial': status = 'trial'
        else: status = 'active'
        paid = d.execute("SELECT COALESCE(SUM(amount),0) s FROM wpayments WHERE workspace_id=?", (w['id'],)).fetchone()['s'] or 0
        # last activity = newest task / update / login in this workspace
        la_t = d.execute("SELECT MAX(created_at) m FROM tasks WHERE workspace_id=?", (w['id'],)).fetchone()['m']
        la_u = d.execute("SELECT MAX(created_at) m FROM updates WHERE workspace_id=?", (w['id'],)).fetchone()['m']
        la_k = d.execute("""SELECT MAX(k.created_at) m FROM tokens k JOIN users u ON u.id=k.user_id
                            WHERE u.workspace_id=?""", (w['id'],)).fetchone()['m']
        last_active = max([x for x in (la_t, la_u, la_k) if x], default=None)
        tasks_week = d.execute("SELECT COUNT(*) n FROM tasks WHERE workspace_id=? AND created_at>=?",
                               (w['id'], (today_ist() - timedelta(days=7)).isoformat())).fetchone()['n']
        # trial progress: Day X of 14
        trial_day = None
        if (w['plan'] or '') == 'Trial' and exp:
            trial_day = max(1, min(14, 14 - (days_left if days_left is not None else 0)))
        # referral info for this workspace
        rby = d.execute("""SELECT w2.name FROM referrals r JOIN workspaces w2 ON w2.id=r.referrer_ws
                           WHERE r.referred_ws=? LIMIT 1""", (w['id'],)).fetchone()
        bonus = d.execute("SELECT COALESCE(SUM(reward_days),0) s FROM referrals WHERE referrer_ws=? AND status='rewarded'",
                          (w['id'],)).fetchone()['s'] or 0
        out.append(dict(id=w['id'], name=w['name'], plan=w['plan'] or 'Pro',
                        expires_on=exp, days_left=days_left, status=status, notes=w['mnotes'] or '',
                        owner_name=owner['name'] if owner else '—', owner_email=owner['email'] if owner else '—',
                        owner_phone='', members=members, clients=clients, tasks_month=tasks_m,
                        tasks_week=tasks_week, last_active=last_active, trial_day=trial_day,
                        referred_by=rby['name'] if rby else None, ref_bonus_days=bonus,
                        creds=_ws_creds(w, owner),
                        protected=ws_protected(w['id']), total_paid=round(paid, 2)))
    rev_month = d.execute("SELECT COALESCE(SUM(amount),0) s FROM wpayments WHERE substr(paid_on,1,7)=?", (month,)).fetchone()['s'] or 0
    new_month = len([x for x in wss if (x['created_at'] or '')[:7] == month])
    renewals_month = d.execute("SELECT COUNT(*) n FROM wpayments WHERE substr(paid_on,1,7)=?", (month,)).fetchone()['n']
    plans = [dict(r) for r in d.execute("SELECT * FROM plans ORDER BY price_m").fetchall()]
    return jsonify(workspaces=out, plans=plans,
                   stats=dict(total=len(out),
                              active=len([w for w in out if w['status'] in ('active', 'trial')]),
                              expiring=len([w for w in out if w['status'] == 'expiring']),
                              expired=len([w for w in out if w['status'] in ('expired', 'suspended')]),
                              users=sum(w['members'] for w in out),
                              new_month=new_month, renewals_month=renewals_month,
                              revenue_month=round(rev_month, 2)))

@app.post('/api/master/workspaces')
@master_required
def master_create_ws():
    d = request.json or {}
    name = (d.get('name') or '').strip()
    oname = (d.get('owner_name') or '').strip()
    email = (d.get('owner_email') or '').strip().lower()
    pw = d.get('password') or ''
    plan = (d.get('plan') or 'Trial').strip()
    months = int(d.get('months') or 0)
    if not name or not oname or not email or len(pw) < 4:
        return jsonify(error="Workspace name, owner name, email & password (4+) required"), 400
    code = (d.get('invite_code') or '').strip() or secrets.token_hex(3).upper()
    if plan == 'Trial':
        exp = (today_ist() + timedelta(days=14)).isoformat()
    elif months > 0:
        exp = (today_ist() + timedelta(days=30 * months)).isoformat()
    else:
        # use the plan's custom package duration if it has one
        pr = db().execute("SELECT duration_val, duration_unit FROM plans WHERE name=?", (plan,)).fetchone()
        if pr and (pr['duration_val'] or 0) > 0:
            exp = (today_ist() + timedelta(days=pr['duration_val'] * _unit_days(pr['duration_unit']))).isoformat()
        else:
            exp = None
    cval = int(d.get('custom_value') or 0)
    cunit = (d.get('custom_unit') or 'days')
    if cval > 0 and cunit in ('days', 'months', 'years'):
        exp = (today_ist() + timedelta(days=cval * _unit_days(cunit))).isoformat()
    creds = json.dumps(dict(email=email, password=pw, invite_code=code))
    cur = db().execute("INSERT INTO workspaces(name,invite_code,created_at,plan,expires_on,wstatus,creds) VALUES(?,?,?,?,?,'active',?)",
                       (name, code, now_ist().isoformat(), plan, exp, creds))
    ws = cur.lastrowid
    salt = secrets.token_hex(8)
    try:
        db().execute("""INSERT INTO users(workspace_id,name,email,salt,pw,role,title,created_at)
                        VALUES(?,?,?,?,?,'owner','Owner',?)""",
                     (ws, oname, email, salt, hashpw(pw, salt), now_ist().isoformat()))
        db().commit()
    except Exception as ex:
        if not _is_unique_violation(ex): raise
        db().execute("DELETE FROM workspaces WHERE id=?", (ws,)); db().commit()
        return jsonify(error="That email is already registered"), 400
    ref = int(d.get('referred_by') or 0)
    rcode = (d.get('ref_code') or '').strip().upper()
    if not ref and rcode:
        rw = db().execute("SELECT id FROM workspaces WHERE UPPER(ref_code)=?", (rcode,)).fetchone()
        if rw: ref = rw['id']
    if ref:
        db().execute("""INSERT INTO referrals(referrer_ws,referred_ws,referred_name,status,created_at)
                        VALUES(?,?,?,'pending',?)""", (ref, ws, name, now_ist().isoformat()))
        db().commit()
    cprice = float(d.get('custom_price') or 0)
    if cprice > 0:
        db().execute("INSERT INTO wpayments(workspace_id,ws_label,amount,note,paid_on) VALUES(?,?,?,?,?)",
                     (ws, '', cprice, 'On workspace creation', today_ist().isoformat()))
        db().commit()
    return jsonify(id=ws, invite_code=code, expires_on=exp, referred_by=ref or None)

@app.post('/api/master/workspaces/<int:wid>/extend')
@master_required
def master_extend(wid):
    d = request.json or {}
    # flexible duration: {value: N, unit: days|months|years}  (backward-compat: {months: N})
    if d.get('months') and not d.get('value'):
        value, unit = int(d['months']), 'months'
    else:
        value, unit = int(d.get('value') or 0), (d.get('unit') or 'days')
    if value <= 0: return jsonify(error="duration required"), 400
    if unit not in ('days', 'months', 'years'): return jsonify(error="bad unit"), 400
    days = value * _unit_days(unit)
    w = db().execute("SELECT expires_on FROM workspaces WHERE id=?", (wid,)).fetchone()
    if not w: return jsonify(error="not found"), 404
    base = today_ist()
    if w['expires_on'] and date.fromisoformat(w['expires_on']) > base:
        base = date.fromisoformat(w['expires_on'])
    new_exp = (base + timedelta(days=days)).isoformat()
    db().execute("UPDATE workspaces SET expires_on=?, wstatus='active' WHERE id=?", (new_exp, wid))
    amount = float(d.get('amount') or 0)
    if amount > 0:
        db().execute("INSERT INTO wpayments(workspace_id,amount,note,paid_on) VALUES(?,?,?,?)",
                     (wid, amount, d.get('note') or f'{value} {unit} extension', today_ist().isoformat()))
    db().commit()
    return jsonify(ok=True, expires_on=new_exp)

@app.get('/api/master/payments')
@master_required
def master_payments():
    d = db()
    rows = d.execute("""SELECT p.id, p.workspace_id, p.ws_label, p.amount, p.note, p.paid_on, w.name AS ws_name
                        FROM wpayments p LEFT JOIN workspaces w ON w.id=p.workspace_id
                        ORDER BY p.paid_on DESC, p.id DESC""").fetchall()
    out, months = [], {}
    for r in rows:
        rec = dict(r)
        rec['ws_name'] = rec['ws_name'] or rec.get('ws_label') or '(deleted workspace)'
        out.append(rec)
        mk = (r['paid_on'] or '')[:7]
        months[mk] = round(months.get(mk, 0) + (r['amount'] or 0), 2)
    year = today_ist().strftime('%Y')
    total_year = round(sum(v for k, v in months.items() if k.startswith(year)), 2)
    total_all = round(sum(months.values()), 2)
    return jsonify(payments=out, month_totals=months, total_year=total_year, total_all=total_all)

@app.post('/api/master/payments')
@master_required
def master_add_payment():
    d = request.json or {}
    amount = float(d.get('amount') or 0)
    if amount <= 0: return jsonify(error="amount required"), 400
    wid = int(d.get('workspace_id') or 0) or None
    label = (d.get('ws_label') or '').strip()
    if not wid and not label: return jsonify(error="pick a workspace or type a name"), 400
    db().execute("INSERT INTO wpayments(workspace_id,ws_label,amount,note,paid_on) VALUES(?,?,?,?,?)",
                 (wid, label, amount, d.get('note', ''), d.get('paid_on') or today_ist().isoformat()))
    db().commit()
    return jsonify(ok=True)

@app.delete('/api/master/payments/<int:pid>')
@master_required
def master_del_payment(pid):
    db().execute("DELETE FROM wpayments WHERE id=?", (pid,)); db().commit()
    return jsonify(ok=True)

@app.put('/api/master/workspaces/<int:wid>')
@master_required
def master_edit_ws(wid):
    d = request.json or {}
    w = db().execute("SELECT * FROM workspaces WHERE id=?", (wid,)).fetchone()
    if not w: return jsonify(error="not found"), 404
    if d.get('wstatus') == 'suspended' and ws_protected(wid):
        return jsonify(error="This is YOUR workspace — it cannot be suspended."), 403
    db().execute("UPDATE workspaces SET plan=?, wstatus=?, mnotes=?, expires_on=? WHERE id=?",
                 (d.get('plan', w['plan']), d.get('wstatus', w['wstatus'] or 'active'),
                  d.get('mnotes', w['mnotes'] or ''), d.get('expires_on', w['expires_on']), wid))
    db().commit()
    return jsonify(ok=True)

@app.post('/api/master/workspaces/<int:wid>/reset-owner')
@master_required
def master_reset_owner(wid):
    pw = (request.json or {}).get('password') or ''
    if len(pw) < 4: return jsonify(error="password min 4 chars"), 400
    u = db().execute("SELECT id FROM users WHERE workspace_id=? AND role='owner' LIMIT 1", (wid,)).fetchone()
    if not u: return jsonify(error="owner not found"), 404
    salt = secrets.token_hex(8)
    db().execute("UPDATE users SET salt=?, pw=? WHERE id=?", (salt, hashpw(pw, salt), u['id']))
    db().execute("DELETE FROM tokens WHERE user_id=?", (u['id'],))
    w = db().execute("SELECT creds FROM workspaces WHERE id=?", (wid,)).fetchone()
    try: c = json.loads(w['creds']) if w and w['creds'] else {}
    except Exception: c = {}
    c['password'] = pw
    db().execute("UPDATE workspaces SET creds=? WHERE id=?", (json.dumps(c), wid))
    db().commit()
    return jsonify(ok=True)

@app.post('/api/master/workspaces/<int:wid>/impersonate')
@master_required
def master_impersonate(wid):
    u = db().execute("SELECT * FROM users WHERE workspace_id=? AND role='owner' LIMIT 1", (wid,)).fetchone()
    if not u: return jsonify(error="owner not found"), 404
    return jsonify(token=issue_token(u['id']), user=userdict(u))

@app.delete('/api/master/workspaces/<int:wid>')
@master_required
def master_delete_ws(wid):
    if ws_protected(wid):
        return jsonify(error="This is YOUR workspace (Buzz Media Fame) — it cannot be deleted."), 403
    db().execute("DELETE FROM workspaces WHERE id=?", (wid,)); db().commit()
    return jsonify(ok=True)

@app.put('/api/master/plans/<int:pid>')
@master_required
def master_edit_plan(pid):
    d = request.json or {}
    p = db().execute("SELECT * FROM plans WHERE id=?", (pid,)).fetchone()
    if not p: return jsonify(error="not found"), 404
    unit = (d.get('duration_unit') or p['duration_unit'] or 'days').lower()
    if unit not in ('days', 'months', 'years'): unit = 'days'
    db().execute("""UPDATE plans SET name=?, price_m=?, price_y=?, max_members=?,
                    duration_val=?, duration_unit=?, price_pack=? WHERE id=?""",
                 ((d.get('name') or p['name']).strip()[:30], int(d.get('price_m') or 0),
                  int(d.get('price_y') or 0), int(d.get('max_members') or 0),
                  int(d.get('duration_val') or 0), unit, int(d.get('price_pack') or 0), pid))
    db().commit()
    return jsonify(ok=True)

@app.post('/api/master/plans')
@master_required
def master_add_plan():
    d = request.json or {}
    if not (d.get('name') or '').strip(): return jsonify(error="name required"), 400
    unit = (d.get('duration_unit') or 'days').lower()
    if unit not in ('days', 'months', 'years'): unit = 'days'
    cur = db().execute("""INSERT INTO plans(name,price_m,price_y,max_members,duration_val,duration_unit,price_pack)
                          VALUES(?,?,?,?,?,?,?)""",
                       (d['name'].strip()[:30], int(d.get('price_m') or 0),
                        int(d.get('price_y') or 0), int(d.get('max_members') or 0),
                        int(d.get('duration_val') or 0), unit, int(d.get('price_pack') or 0)))
    db().commit()
    return jsonify(id=cur.lastrowid)

@app.get('/api/master/settings')
@master_required
def master_get_settings():
    rows = db().execute("SELECT key, value FROM msettings").fetchall()
    return jsonify({r['key']: r['value'] for r in rows})

@app.put('/api/master/settings')
@master_required
def master_put_settings():
    d = request.json or {}
    for k, v in d.items():
        db().execute("""INSERT INTO msettings(key,value) VALUES(?,?)
                        ON CONFLICT(key) DO UPDATE SET value=?""", (str(k)[:50], str(v)[:2000], str(v)[:2000]))
    db().commit()
    return jsonify(ok=True)

@app.delete('/api/master/plans/<int:pid>')
@master_required
def master_del_plan(pid):
    db().execute("DELETE FROM plans WHERE id=?", (pid,)); db().commit()
    return jsonify(ok=True)

# ---------------- Refer & Earn (owner side) ----------------
def get_ref_code(ws):
    w = db().execute("SELECT ref_code, name FROM workspaces WHERE id=?", (ws,)).fetchone()
    if w['ref_code']: return w['ref_code']
    base = ''.join(c for c in (w['name'] or 'BF').upper() if c.isalnum())[:4] or 'BF'
    code = base + '-' + secrets.token_hex(2).upper()
    db().execute("UPDATE workspaces SET ref_code=? WHERE id=?", (code, ws)); db().commit()
    return code

@app.get('/api/referral')
@role_required('owner')
def referral_info():
    ws = wsid()
    code = get_ref_code(ws)
    rows = db().execute("""SELECT referred_name, status, reward_note, created_at, rewarded_at
                           FROM referrals WHERE referrer_ws=? ORDER BY id DESC""", (ws,)).fetchall()
    return jsonify(code=code, referrals=[dict(r) for r in rows])

# ---------------- Master: referrals ----------------
@app.get('/api/master/referrals')
@master_required
def master_referrals():
    rows = db().execute("""SELECT r.*, w1.name AS referrer_name, w2.name AS referred_ws_name
                           FROM referrals r
                           JOIN workspaces w1 ON w1.id=r.referrer_ws
                           LEFT JOIN workspaces w2 ON w2.id=r.referred_ws
                           ORDER BY r.id DESC""").fetchall()
    codes = db().execute("SELECT id, name, ref_code FROM workspaces WHERE ref_code IS NOT NULL").fetchall()
    return jsonify(referrals=[dict(r) for r in rows], codes=[dict(c) for c in codes])

@app.post('/api/master/referrals')
@master_required
def master_add_referral():
    d = request.json or {}
    rid = int(d.get('referrer_ws') or 0)
    if not rid: return jsonify(error="referrer workspace required"), 400
    val = int(d.get('value') or 0)
    unit = (d.get('unit') or 'months')
    if unit not in ('days', 'months', 'years'): unit = 'months'
    disc = int(d.get('discount') or 0)
    days = val * _unit_days(unit) if val > 0 else 0
    parts = []
    if val: parts.append(f'plan: {val} {unit} free')
    if disc: parts.append(f'{disc}% off for referred')
    db().execute("""INSERT INTO referrals(referrer_ws,referred_ws,referred_name,status,reward_note,reward_days,created_at)
                    VALUES(?,?,?,'pending',?,?,?)""",
                 (rid, d.get('referred_ws') or None, (d.get('referred_name') or '').strip(),
                  ' · '.join(parts), days, now_ist().isoformat()))
    db().commit()
    return jsonify(ok=True)

@app.post('/api/master/referrals/<int:rid>/reward')
@master_required
def master_reward_referral(rid):
    d = request.json or {}
    r = db().execute("SELECT * FROM referrals WHERE id=?", (rid,)).fetchone()
    if not r: return jsonify(error="not found"), 404
    # custom duration: value + unit (days/months/years); legacy: days
    if d.get('value'):
        value, unit = int(d['value']), (d.get('unit') or 'days')
        if unit not in ('days', 'months', 'years'): return jsonify(error="bad unit"), 400
        days = value * _unit_days(unit)
        note = f'+{value} {unit} free'
    else:
        days = int(d.get('days') or 0) or (r['reward_days'] or 30)
        note = f'+{days} days free'
    new_exp = None
    if not ws_protected(r['referrer_ws']):
        w = db().execute("SELECT expires_on FROM workspaces WHERE id=?", (r['referrer_ws'],)).fetchone()
        base = today_ist()
        if w and w['expires_on'] and date.fromisoformat(w['expires_on']) > base:
            base = date.fromisoformat(w['expires_on'])
        new_exp = (base + timedelta(days=days)).isoformat()
        db().execute("UPDATE workspaces SET expires_on=? WHERE id=?", (new_exp, r['referrer_ws']))
    db().execute("UPDATE referrals SET status='rewarded', reward_note=?, reward_days=?, rewarded_at=? WHERE id=?",
                 (note, days, now_ist().isoformat(), rid))
    db().commit()
    return jsonify(ok=True, new_expiry=new_exp)

@app.put('/api/master/referrals/<int:rid>')
@master_required
def master_edit_referral(rid):
    d = request.json or {}
    r = db().execute("SELECT * FROM referrals WHERE id=?", (rid,)).fetchone()
    if not r: return jsonify(error="not found"), 404
    db().execute("""UPDATE referrals SET referrer_ws=?, referred_ws=?, referred_name=?, status=? WHERE id=?""",
                 (int(d.get('referrer_ws') or r['referrer_ws']),
                  d.get('referred_ws') if d.get('referred_ws') else r['referred_ws'],
                  d.get('referred_name', r['referred_name']),
                  d.get('status', r['status']), rid))
    db().commit()
    return jsonify(ok=True)

@app.post('/api/master/reset-settings')
@master_required
def master_reset_settings():
    db().execute("DELETE FROM msettings"); db().commit()
    return jsonify(ok=True)

@app.get('/api/master/refcode/<code>')
@master_required
def master_refcode(code):
    w = db().execute("SELECT id, name FROM workspaces WHERE UPPER(ref_code)=?", (code.strip().upper(),)).fetchone()
    if not w: return jsonify(error="Unknown referral code"), 404
    return jsonify(id=w['id'], name=w['name'])

@app.delete('/api/master/referrals/<int:rid>')
@master_required
def master_del_referral(rid):
    db().execute("DELETE FROM referrals WHERE id=?", (rid,)); db().commit()
    return jsonify(ok=True)

# ---------------- Master: calendar (day-wise events incl. money) ----------------
@app.get('/api/master/calendar')
@master_required
def master_calendar():
    month = request.args.get('month') or today_ist().strftime('%Y-%m')
    d = db()
    events = {}
    def add(day, kind, text, amount=0):
        if not day or not day.startswith(month): return
        events.setdefault(day, []).append(dict(kind=kind, text=text, amount=amount))
    for p in d.execute("""SELECT p.amount, p.note, p.paid_on, w.name FROM wpayments p
                          LEFT JOIN workspaces w ON w.id=p.workspace_id""").fetchall():
        add(p['paid_on'], 'payment', f"{p['name'] or 'Deleted ws'} paid" + (f" · {p['note']}" if p['note'] else ''), p['amount'])
    for w in d.execute("SELECT name, created_at, expires_on FROM workspaces").fetchall():
        add((w['created_at'] or '')[:10], 'created', f"Workspace created: {w['name']}")
        add(w['expires_on'], 'expiry', f"Expires: {w['name']}")
    for r in d.execute("""SELECT r.created_at, r.rewarded_at, r.referred_name, w.name
                          FROM referrals r JOIN workspaces w ON w.id=r.referrer_ws""").fetchall():
        add((r['created_at'] or '')[:10], 'referral', f"Referral by {r['name']}: {r['referred_name'] or 'new lead'}")
        add((r['rewarded_at'] or '')[:10], 'reward', f"Reward given to {r['name']}")
    day_money = {k: round(sum(e['amount'] for e in v if e['kind'] == 'payment'), 2) for k, v in events.items()}
    total = round(sum(day_money.values()), 2)
    return jsonify(month=month, events=events, day_money=day_money, month_total=total)

# ---------------- Guest / client read-only link ----------------
@app.post('/api/clients/<int:cid>/guest-link')
@role_required('owner', 'admin')
def guest_link(cid):
    c = own_client(cid)
    if not c: return jsonify(error="not found"), 404
    tok = c['guest_token']
    if not tok or (request.json or {}).get('regenerate'):
        tok = secrets.token_urlsafe(16)
        db().execute("UPDATE clients SET guest_token=? WHERE id=?", (tok, cid)); db().commit()
    return jsonify(token=tok, url=f"/client/{tok}")

@app.delete('/api/clients/<int:cid>/guest-link')
@role_required('owner', 'admin')
def guest_link_off(cid):
    if not own_client(cid): return jsonify(error="not found"), 404
    db().execute("UPDATE clients SET guest_token=NULL WHERE id=?", (cid,)); db().commit()
    return jsonify(ok=True)

@app.get('/api/guest/<token>')
def guest_data(token):
    c = db().execute("SELECT * FROM clients WHERE guest_token=?", (token,)).fetchone()
    if not c: return jsonify(error="Invalid or expired link"), 404
    if ws_locked(c['workspace_id']): return jsonify(error="This dashboard is currently unavailable"), 403
    d = db(); month = today_ist().strftime('%Y-%m')
    w = d.execute("SELECT name FROM workspaces WHERE id=?", (c['workspace_id'],)).fetchone()
    raw = get_ws_setting(c['workspace_id'], 'company')
    agency = w['name']
    try:
        comp = json.loads(raw) if raw else {}
        agency = comp.get('name') or agency
    except Exception: pass
    qs = d.execute("SELECT service, monthly_target FROM quotas WHERE client_id=?", (c['id'],)).fetchall()
    quotas = []
    for q in qs:
        done = d.execute("""SELECT COALESCE(SUM(qty),0) n FROM tasks WHERE client_id=? AND service=?
            AND status='done' AND substr(done_at,1,7)=?""", (c['id'], q['service'], month)).fetchone()['n']
        quotas.append(dict(service=q['service'], target=q['monthly_target'], done=done))
    # extra services delivered without quota
    quota_svcs = {q['service'] for q in qs}
    extra = d.execute("""SELECT service, COALESCE(SUM(qty),0) n FROM tasks WHERE client_id=?
        AND status='done' AND substr(done_at,1,7)=? GROUP BY service""", (c['id'], month)).fetchall()
    for e in extra:
        if e['service'] not in quota_svcs and e['n'] > 0:
            quotas.append(dict(service=e['service'], target=0, done=e['n']))
    tasks = d.execute("""SELECT title, service, qty, status, due, done_at FROM tasks
                         WHERE client_id=? AND (substr(created_at,1,7)=? OR substr(done_at,1,7)=? OR status!='done')
                         ORDER BY CASE WHEN due IS NULL THEN 1 ELSE 0 END, due""",
                      (c['id'], month, month)).fetchall()
    cal = {}
    for t in tasks:
        key = (t['done_at'] or '')[:10] if t['status'] == 'done' else (t['due'] or '')[:10]
        if key: cal.setdefault(key, []).append(dict(title=t['title'], service=t['service'], status=t['status'], qty=t['qty']))
    done_n = len([t for t in tasks if t['status'] == 'done'])
    return jsonify(client=c['name'], agency=agency, month=today_ist().strftime('%B %Y'),
                   quotas=quotas, calendar=cal,
                   stats=dict(total=len(tasks), done=done_n, in_progress=len(tasks) - done_n))

@app.get('/client/<token>')
def guest_page(token):
    return send_from_directory('static', 'guest.html')

# public pricing feed for the marketing website (no auth, CORS open)
@app.get('/api/public/plans')
def public_plans():
    seed_plans()
    plans = [dict(r) for r in db().execute("SELECT name, price_m, price_y, max_members, duration_val, duration_unit, price_pack FROM plans ORDER BY price_m").fetchall()]
    resp = jsonify(plans=plans)
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@app.get('/master')
def master_page():
    return send_from_directory('static', 'master.html')

@app.get('/api/meta')
def meta():
    return jsonify(services=SERVICE_TYPES)

@app.get('/')
def index():
    return send_from_directory('static', 'index.html')

init_db()
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
